#import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


# Criando uma workbook

wb = Workbook() # criando objeto 
ws = wb.active # Planilha ativa 
ws.title = 'Dados_1' # Nomeando a planilha ativa
nome_arquivo = 'teste_1.xlsx' 
wb.save(nome_arquivo) # Salvado a workbook em um arquivo excel (xlsx)

# Lendo os dados do arquivo excel

wb_1 = load_workbook(nome_arquivo, )




